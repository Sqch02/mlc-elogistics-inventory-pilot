#!/usr/bin/env python3
"""
Script to parse FLORNA - SUIVI RECLAMATION.xlsx and generate SQL for claims import.
"""

import openpyxl
from datetime import datetime
import json

TENANT_ID = "00000000-0000-0000-0000-000000000001"
FILE_PATH = "docs/tableau_client/FLORNA - SUIVI RECLAMATION.xlsx"

# Month columns structure (each block has 6 columns)
# Format: (start_col, year, month) - start_col is 1-indexed (openpyxl style)
# Columns: DATE | N° DE COMMANDE | RÉEXPEDIÉE | PAR QUI | TRACKING | COMMENTAIRES
MONTH_BLOCKS = [
    (2, 2025, 12),   # DÉCEMBRE 2025 (B-G)
    (8, 2025, 11),   # NOVEMBRE 2025 (H-M)
    (14, 2025, 10),  # OCTOBRE 2025 (N-S)
    (20, 2025, 9),   # SEPTEMBRE 2025 (T-Y)
    (26, 2025, 8),   # AOUT 2025 (Z-AE)
]

def parse_date(date_val, year, month):
    """Parse date from various formats."""
    if not date_val:
        return None
    try:
        # Handle datetime objects (openpyxl often returns these)
        if hasattr(date_val, 'strftime'):
            return date_val.strftime('%Y-%m-%d')

        # Handle string formats like "2025-12-01 00:0..."
        date_str = str(date_val)
        if date_str.startswith('20') and '-' in date_str:
            # Already in YYYY-MM-DD format
            return date_str[:10]

        # Handle formats like "01/12", "02/11", etc.
        if "/" in date_str:
            parts = date_str.split("/")
            day = int(parts[0])
            return f"{year}-{month:02d}-{day:02d}"
        elif isinstance(date_val, (int, float)):
            day = int(date_val)
            return f"{year}-{month:02d}-{day:02d}"
    except:
        pass
    return None

def clean_string(s):
    """Clean string for SQL insertion."""
    if s is None:
        return None
    s = str(s).strip()
    if s in ['', 'None', 'nan', '/']:
        return None
    # Escape single quotes for SQL
    s = s.replace("'", "''")
    # Remove leading/trailing whitespace and tab characters
    s = ' '.join(s.split())
    return s

def clean_order_ref(val):
    """Clean order reference - remove .0 from floats."""
    if val is None:
        return None
    s = str(val).strip()
    if s in ['', 'None', 'nan', '/']:
        return None
    # Remove .0 suffix from float values
    if s.endswith('.0'):
        s = s[:-2]
    return s

def parse_excel():
    """Parse the Excel file and extract claims data."""
    wb = openpyxl.load_workbook(FILE_PATH, data_only=True)
    ws = wb.active

    claims = []

    # Data starts at row 5 (1-indexed, row 4 is headers)
    for row_idx in range(5, ws.max_row + 1):
        for start_col, year, month in MONTH_BLOCKS:
            # Column indices (1-indexed for openpyxl: A=1, B=2, etc.)
            date_col = start_col + 0      # DATE
            order_col = start_col + 1     # N° DE COMMANDE
            reshipped_col = start_col + 2 # COMMANDE RÉEXPEDIÉE
            by_whom_col = start_col + 3   # PAR QUI ?
            tracking_col = start_col + 4  # NUMÉRO DE SUIVI
            comment_col = start_col + 5   # COMMENTAIRES

            date_val = ws.cell(row=row_idx, column=date_col).value
            order_val = ws.cell(row=row_idx, column=order_col).value
            reshipped_val = ws.cell(row=row_idx, column=reshipped_col).value
            by_whom_val = ws.cell(row=row_idx, column=by_whom_col).value
            tracking_val = ws.cell(row=row_idx, column=tracking_col).value
            comment_val = ws.cell(row=row_idx, column=comment_col).value

            # Skip empty rows or non-reshipped
            if not order_val:
                continue

            # Check if reshipped (TRUE, True, 1, etc.)
            is_reshipped = str(reshipped_val).upper() in ['TRUE', '1', 'OUI', 'YES']
            if not is_reshipped:
                continue

            # Parse date
            opened_at = parse_date(date_val, year, month)
            if not opened_at:
                continue

            # Clean order ref
            order_ref = clean_order_ref(order_val)
            if not order_ref:
                continue

            # Build description
            desc_parts = []
            if by_whom_val and clean_string(by_whom_val):
                desc_parts.append(f"Par {clean_string(by_whom_val)}")
            if comment_val and clean_string(comment_val):
                desc_parts.append(clean_string(comment_val))
            description = ". ".join(desc_parts) if desc_parts else "Réexpédition"

            # Tracking as decision note (clean .0 suffix from numbers)
            decision_note = None
            if tracking_val:
                tracking_clean = clean_order_ref(tracking_val)  # Remove .0 suffix
                if tracking_clean:
                    decision_note = f"Tracking: {tracking_clean}"

            claims.append({
                'order_ref': order_ref,
                'opened_at': opened_at,
                'description': description,
                'decision_note': decision_note,
            })

    return claims

def generate_sql(claims):
    """Generate SQL statements for claims import."""
    sql_statements = []

    for claim in claims:
        desc = f"'{claim['description']}'" if claim['description'] else 'NULL'
        note = f"'{claim['decision_note']}'" if claim['decision_note'] else 'NULL'

        sql = f"""INSERT INTO claims (tenant_id, order_ref, opened_at, status, description, decision_note)
SELECT
  '{TENANT_ID}',
  '{claim['order_ref']}',
  '{claim['opened_at']}'::date,
  'cloturee'::claim_status,
  {desc},
  {note}
WHERE NOT EXISTS (
  SELECT 1 FROM claims
  WHERE tenant_id = '{TENANT_ID}'
  AND order_ref = '{claim['order_ref']}'
);"""
        sql_statements.append(sql)

    return sql_statements

def main():
    print("Parsing Excel file...")
    claims = parse_excel()
    print(f"Found {len(claims)} claims to import")

    # Remove duplicates by order_ref (keep first occurrence)
    seen = set()
    unique_claims = []
    for claim in claims:
        if claim['order_ref'] not in seen:
            seen.add(claim['order_ref'])
            unique_claims.append(claim)

    print(f"After deduplication: {len(unique_claims)} unique claims")

    # Generate SQL
    sql_statements = generate_sql(unique_claims)

    # Output combined SQL
    combined_sql = "\n\n".join(sql_statements)

    # Save to file
    with open("scripts/claims_import.sql", "w") as f:
        f.write(f"-- Claims import generated on {datetime.now()}\n")
        f.write(f"-- Total claims: {len(unique_claims)}\n\n")
        f.write(combined_sql)

    print(f"SQL saved to scripts/claims_import.sql")

    # Also output JSON for verification
    with open("scripts/claims_data.json", "w") as f:
        json.dump(unique_claims, f, indent=2, default=str)

    print(f"JSON data saved to scripts/claims_data.json")

    # Print first 5 for verification
    print("\nFirst 5 claims:")
    for claim in unique_claims[:5]:
        print(f"  - {claim['order_ref']} ({claim['opened_at']}): {claim['description'][:50]}...")

if __name__ == "__main__":
    main()
